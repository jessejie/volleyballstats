# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.

import sys
import re
from openpyxl import load_workbook

SEPARATOR1 = "=========="
SEPARATOR2 = "----------"


def perct(a, b, decimals=0):
    if b == 0:
        return "N/A"
    # round has some formatting issues when specifying 0 digits
    return f"{round(a * 100.0 / b)}" if decimals == 0 else f"{round(a * 100.0 / b, decimals)}"

# Volleyball stats following this legend:
# H	  -> Hit but opponents get offense
# HF  -> Hit causing freeball
# K   -> Kill
# HE  -> Hitting Error
# P3  -> 3 option pass
# P2  -> 2 option pass
# P1  -> 1 option pass
# P0  -> 0 option pass (shank or overpass)
# S3  -> Serve leading to P3
# S2  -> Serve leading to P2
# S1  -> Serve leading to P1
# S0  -> Serve leading to P0
# SE  -> Serve Error
# B   -> Block
# BE  -> Block Error
# BK  -> Block Kill
# D   -> Defense
# DE  -> Defense error
# F   -> Freeball pass
# FE  -> Freeball error
# SeE -> Setting Error
# CE  -> Communication Error
# C   -> Cover


class Player:
    hits = 0
    hit_free = 0
    kills = 0
    hitting_errors = 0
    p3 = 0
    p2 = 0
    p1 = 0
    p0 = 0
    s3 = 0
    s2 = 0
    s1 = 0
    s0 = 0
    serve_error = 0
    block = 0
    block_error = 0
    block_kill = 0
    defense = 0
    defense_error = 0
    freeball = 0
    freeball_error = 0
    setting_error = 0
    comm_error = 0
    cover = 0
    total_sets = 0

    def __init__(self, name, is_setter=False):
        self.name = name
        self.is_setter = is_setter
        self.hitters = []

    def get_total_hits(self):
        return self.hits + self.hit_free + self.kills + self.hitting_errors

    def get_total_passes(self):
        return self.p3 + self.p2 + self.p1 + self.p0

    def get_total_serves(self):
        return self.s3 + self.s2 + self.s1 + self.s0 + self.serve_error

    def get_total_defense(self):
        return self.defense + self.defense_error

    def get_total_freeball(self):
        return self.freeball + self.freeball_error

    def get_total_blocks(self):
        return self.block + self.block_kill + self.block_error

    def generate_report(self):
        total_hits = self.get_total_hits()
        total_passes = self.get_total_passes()
        total_serves = self.get_total_serves()
        total_defense = self.get_total_defense()
        total_freeball = self.get_total_freeball()
        total_blocks = self.get_total_blocks()

        report = f"""
{self.name}
{SEPARATOR1}
HITTING
{SEPARATOR2}
Total Hits: {total_hits}
Effective Hit Rating: {perct(self.kills + (0.5 * self.hit_free) - self.hitting_errors, total_hits)}
Kill Pct: {perct(self.kills, total_hits)}%
Hit Freeball Pct: {perct(self.hit_free, total_hits)}%
Hit with Opponent Offense Pct: {perct(self.hits, total_hits)}%
Error Rate: {perct(self.hitting_errors, total_hits)}%

PASSING
{SEPARATOR2}
Total Passes: {total_passes}
Effective Pass Rating: {perct(self.p3 + self.p2 - (0.5 * self.p1) - self.p0, total_passes)}
P3 Pct: {perct(self.p3, total_passes)}%
P2 Pct: {perct(self.p2, total_passes)}%
P1 Pct: {perct(self.p1, total_passes)}%
P0 Pct: {perct(self.p0, total_passes)}%
Avg Pass: {perct(self.p3 * 3 + self.p2 * 2 + self.p1, total_passes * 100, decimals=2)}
Total Freeballs: {total_freeball}
Freeball Pct: {perct(self.freeball, total_freeball)}%

SERVING
{SEPARATOR2}
Total Serves: {total_serves}
Effective Serve Rating: {perct(self.s0 + self.s1 - (0.5 * self.s3) - self.serve_error, total_serves)}
S0 Pct: {perct(self.s0, total_serves)}%
S1 Pct: {perct(self.s1, total_serves)}%
S2 Pct: {perct(self.s2, total_serves)}%
S3 Pct: {perct(self.s3, total_serves)}%
Error Pct: {perct(self.serve_error, total_serves)}%

DEFENSE
{SEPARATOR2}
Total Defenses: {total_defense}
Defense Pct: {perct(self.defense, total_defense)}%
Covers: {self.cover}
Total Blocks: {total_blocks}
Block Rating: {perct((0.5 * self.block) + self.block_kill - self.block_error, total_blocks)}
Block Kill Pct: {perct(self.block_kill, total_blocks)}%
Block Error Pct: {perct(self.block_error, total_blocks)}%
        """

        if self.comm_error > 0:
            report += f"""
COMMUNICATION
{SEPARATOR2}
Communication Errors: {self.comm_error}
            """

        if self.total_sets > 0:
            report += f"""
SETTING
{SEPARATOR2}
Total Sets: {self.total_sets}
Setting Error Pct: {perct(self.setting_error, self.total_sets)}%
            """
        elif self.setting_error > 0:
            report += f"""
SETTING
{SEPARATOR2}
Setting Errors: {self.setting_error}
            """

        return report


def parse_action(act, setter, player):
    # Only supports a single setter
    action = act.upper()
    if action == "H":
        player.hits += 1
        if not player.is_setter:
            setter.total_sets += 1
    elif action == "HF":
        player.hit_free += 1
        if not player.is_setter:
            setter.total_sets += 1
    elif action == "HE":
        player.hitting_errors += 1
        if not player.is_setter:
            setter.total_sets += 1
    elif action == "K":
        player.kills += 1
        if not player.is_setter:
            setter.total_sets += 1
    elif action == "P3":
        player.p3 += 1
    elif action == "P2":
        player.p2 += 1
    elif action == "P1":
        player.p1 += 1
    elif action == "P0":
        player.p0 += 1
    elif action == "S3":
        player.s3 += 1
    elif action == "S2":
        player.s2 += 1
    elif action == "S1":
        player.s1 += 1
    elif action == "S0":
        player.s0 += 1
    elif action == "SE":
        player.serve_error += 1
    elif action == "B":
        player.block += 1
    elif action == "BK":
        player.block_kill += 1
    elif action == "BE":
        player.block_error += 1
    elif action == "D":
        player.defense += 1
    elif action == "DE":
        player.defense_error += 1
    elif action == "F":
        player.freeball += 1
    elif action == "FE":
        player.freeball_error += 1
    elif action == "SEE":
        player.setting_error += 1
    elif action == "CE":
        player.comm_error += 1
    elif action == "C":
        player.cover += 1


def parse_score_rows(player_index, setter, rows):
    for summary in rows:
        for i in range(1, len(summary)):
            player = player_index[i-1]
            actions_string = summary[i]
            if actions_string:
                actions = actions_string.split(" ")
                for action in actions:
                    parse_action(action, setter, player)


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    if len(sys.argv) != 3:
        raise RuntimeError("Wrong number of arguments. Use `main.py \"spreadsheet.xlsx\" \"W4(.*)\"`")

    file = sys.argv[1]
    workbook = load_workbook(filename=file)

    sheet_regex = sys.argv[2]
    worksheets = [worksheet for worksheet in workbook.worksheets if re.match(sheet_regex, worksheet.title)]

    players = {}

    for worksheet in worksheets:
        print(f"Parsing Worksheet: {worksheet.title}")
        player_index = []
        setter = None
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows)
        for item in header:
            if item != "Score":
                # We only support 5-1
                if "Setter" in item:
                    if item not in players:
                        players[item] = Player(item, is_setter=True)
                    setter = players[item]
                else:
                    if item not in players:
                        players[item] = Player(item)

                player_index.append(players[item])

        parse_score_rows(player_index, setter, rows)

    full_report = ""
    for player in players.values():
        full_report += player.generate_report()

    with open("report.txt", "w") as output:
        output.write(full_report)
